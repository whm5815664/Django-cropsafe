# -*- coding: utf-8 -*-
import json
import math
import os
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent.parent
WORKSPACE = Path(os.environ.get("CROPSAFE_WORKSPACE", PROJECT_DIR))
ROOT = Path(os.environ.get("CROPSAFE_DATA_ROOT", PROJECT_DIR / "data"))
OUTPUT = WORKSPACE / "data.js"

REQUIRED_DIRS = (
    "前10大主产国汇总",
    "进口来源情况",
    "风险分级结果",
    "短期替代潜能",
    "长期替代潜能",
)


CROPS = {
    "soybean": {
        "name": "大豆",
        "subtitle": "基于实验室真实结果呈现主产格局、进口来源、CI 风险分级与替代潜能。",
        "producer": ROOT / "前10大主产国汇总" / "大豆前10主产国汇总(soybean_top10_producers).xlsx",
        "imports": ROOT / "进口来源情况" / "中国大豆进口分析(china_soybean_import_analysis).xlsx",
        "risk": ROOT / "风险分级结果" / "大豆风险分级（soybean_risk_classification）.xlsx",
        "short": ROOT / "短期替代潜能" / "大豆短期替代潜能（soybean_st_sub_potential）.xlsx",
        "long": ROOT / "长期替代潜能" / "大豆长期替代潜能（soybean_lt_sub_potential）.xlsx",
    },
    "corn": {
        "name": "玉米",
        "subtitle": "基于实验室真实结果呈现主产格局、进口来源、CI 风险分级与替代潜能。",
        "producer": ROOT / "前10大主产国汇总" / "玉米前10主产国汇总(maize_top10_producers).xlsx",
        "imports": ROOT / "进口来源情况" / "中国玉米进口分析(china_maize_import_analysis).xlsx",
        "risk": ROOT / "风险分级结果" / "玉米风险分级（maize_risk_classification）.xlsx",
        "short": ROOT / "短期替代潜能" / "玉米短期替代潜能（corn_st_sub_potential）.xlsx",
        "long": ROOT / "长期替代潜能" / "玉米长期替代潜能（maize_lt_sub_potential）.xlsx",
    },
    "wheat": {
        "name": "小麦",
        "subtitle": "基于实验室真实结果呈现主产格局、进口来源、CI 风险分级与替代潜能。",
        "producer": ROOT / "前10大主产国汇总" / "小麦前10主产国汇总(wheat_top10_producers).xlsx",
        "imports": ROOT / "进口来源情况" / "中国小麦进口分析(china_wheat_import_analysis).xlsx",
        "risk": ROOT / "风险分级结果" / "小麦风险分级（wheat_risk_classification）.xlsx",
        "short": ROOT / "短期替代潜能" / "小麦短期替代潜能（wheat_st_sub_potential）.xlsx",
        "long": ROOT / "长期替代潜能" / "小麦长期替代潜能（wheat_lt_sub_potential）.xlsx",
    },
    "barley": {
        "name": "大麦",
        "subtitle": "基于实验室真实结果呈现主产格局、进口来源、CI 风险分级与替代潜能。",
        "producer": ROOT / "前10大主产国汇总" / "大麦前10主产国汇总(barley_top10_producers).xlsx",
        "imports": ROOT / "进口来源情况" / "中国大麦进口分析(china_barley_import_analysis).xlsx",
        "risk": ROOT / "风险分级结果" / "大麦风险分级（barley_risk_classification）.xlsx",
        "short": ROOT / "短期替代潜能" / "大麦短期替代潜能（barley_st_sub_potential）.xlsx",
        "long": ROOT / "长期替代潜能" / "大麦长期替代潜能（barley_lt_sub_potential）.xlsx",
    },
    "sorghum": {
        "name": "高粱",
        "subtitle": "基于实验室真实结果呈现主产格局、进口来源、CI 风险分级与替代潜能。",
        "producer": ROOT / "前10大主产国汇总" / "高粱前10主产国汇总(sorghum_top10_producers).xlsx",
        "imports": ROOT / "进口来源情况" / "中国高粱进口分析(china_sorghum_import_analysis).xlsx",
        "risk": ROOT / "风险分级结果" / "高粱风险分级（sorghum_risk_classification）.xlsx",
        "short": ROOT / "短期替代潜能" / "高粱短期替代潜能（sorghum_st_sub_potential）.xlsx",
        "long": ROOT / "长期替代潜能" / "高粱长期替代潜能（sorghum_lt_sub_potential）.xlsx",
    },
}


COUNTRY_LABELS = {
    "USA": "美国",
    "United States of America": "美国",
    "Brazil": "巴西",
    "Argentina": "阿根廷",
    "Uruguay": "乌拉圭",
    "Canada": "加拿大",
    "Paraguay": "巴拉圭",
    "Russian Federation": "俄罗斯",
    "Russia": "俄罗斯",
    "Ukraine": "乌克兰",
    "Australia": "澳大利亚",
    "France": "法国",
    "Germany": "德国",
    "India": "印度",
    "China": "中国",
    "Myanmar": "缅甸",
    "Mexico": "墨西哥",
    "Nigeria": "尼日利亚",
    "Ethiopia": "埃塞俄比亚",
    "Sudan": "苏丹",
    "Kazakhstan": "哈萨克斯坦",
    "South Africa": "南非",
    "Romania": "罗马尼亚",
    "Bulgaria": "保加利亚",
    "Türkiye": "土耳其",
    "Turkey": "土耳其",
    "Poland": "波兰",
    "Spain": "西班牙",
    "Hungary": "匈牙利",
    "Serbia": "塞尔维亚",
    "Austria": "奥地利",
}


COUNTRY_COORDINATES = {
    "China": [730, 245],
    "USA": [205, 190],
    "United States of America": [205, 190],
    "Brazil": [325, 365],
    "Argentina": [305, 432],
    "Uruguay": [325, 422],
    "Canada": [215, 135],
    "Paraguay": [323, 390],
    "Russian Federation": [670, 122],
    "Ukraine": [565, 180],
    "Myanmar": [755, 274],
    "Australia": [838, 382],
    "France": [505, 188],
    "Germany": [520, 180],
    "India": [705, 236],
    "Mexico": [170, 240],
    "Nigeria": [497, 298],
    "Ethiopia": [585, 292],
    "Sudan": [574, 275],
    "Kazakhstan": [648, 170],
    "South Africa": [560, 430],
    "Romania": [542, 185],
    "Bulgaria": [548, 190],
    "Türkiye": [560, 204],
    "Turkey": [560, 204],
    "Poland": [530, 168],
    "Spain": [480, 214],
    "Hungary": [540, 178],
    "Serbia": [538, 188],
    "Austria": [528, 178],
}


SHORT_TERM_TIER = {"高": 1, "较高": 1, "中": 2, "较低": 3, "低": 3}
LONG_TERM_TIER = {"极高": 1, "高": 1, "较高": 2, "中": 2, "较低": 3, "低": 3}
RISK_CLASS_ORDER = {"very_high": 4, "high": 3, "medium": 2, "low": 1}
RISK_LABELS = {"very_high": "极高风险", "high": "高风险", "medium": "中风险", "low": "低风险"}


def round2(value):
    if value is None:
        return None
    return round(float(value), 4)


def label_country(name):
    return COUNTRY_LABELS.get(name, name)


def load_sheet_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    data_rows = []
    for row in rows[2:]:
        if not row or all(cell is None for cell in row):
            continue
        entry = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            entry[header] = row[index]
        data_rows.append(entry)
    return data_rows


def load_import_history(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["总汇总"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    data = []
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        item = {}
        for index, header in enumerate(headers):
            item[header] = row[index]
        data.append(item)
    return data


def load_producer_history(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    yearly = []
    for ws in wb.worksheets:
        try:
            year = int(str(ws.title).replace("年", ""))
        except ValueError:
            continue
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[1]
        for row in rows[2:]:
            if not row or row[0] is None:
                continue
            item = {"year": year}
            for index, header in enumerate(headers):
                item[header] = row[index]
            yearly.append(item)
    return yearly


def group_years(rows, key):
    years = sorted({int(item[key]) for item in rows if item.get(key) is not None})
    return years


def build_product(crop_id, config):
    producer_rows = load_producer_history(config["producer"])
    import_rows = load_import_history(config["imports"])
    risk_rows = load_sheet_rows(config["risk"])
    short_rows = load_sheet_rows(config["short"])
    long_rows = load_sheet_rows(config["long"])

    years = sorted({int(item["year"]) for item in producer_rows})

    producer_by_country = {}
    for row in producer_rows:
        country = row["country"]
        record = producer_by_country.setdefault(country, {})
        record[int(row["year"])] = row

    producers = {}
    for country, entries in producer_by_country.items():
        producers[country] = {
            "output": [round2(entries.get(year, {}).get("production_10kt", 0) / 100) for year in years],
            "area": [round2(entries.get(year, {}).get("area_10kha", 0) / 100) for year in years],
            "yield": [round2(entries.get(year, {}).get("yield_t_per_ha", 0)) for year in years],
        }

    import_by_country = {}
    for row in import_rows:
        country = row["importer"]
        record = import_by_country.setdefault(country, {})
        record[int(row["year"])] = row

    imports = {}
    for country, entries in import_by_country.items():
        imports[country] = {
            "volume": [round2(entries.get(year, {}).get("import_volume_10kt", 0) / 100) for year in years],
            "price": [round2(entries.get(year, {}).get("price_usd_per_kg", 0) * 1000) for year in years],
            "dependency": [round2(entries.get(year, {}).get("import_concentration_pct", 0) / 100) for year in years],
        }

    production_trend = []
    import_trend = []
    for year in years:
        year_producer_total = sum(
            float(row["production_10kt"] or 0) for row in producer_rows if int(row["year"]) == year
        )
        year_import_total = sum(
            float(row["import_volume_10kt"] or 0) for row in import_rows if int(row["year"]) == year
        )
        production_trend.append(round2(year_producer_total / 100))
        import_trend.append(round2(year_import_total / 100))

    latest_year = years[-1]
    latest_imports = sorted(
        [row for row in import_rows if int(row["year"]) == latest_year],
        key=lambda item: float(item["import_volume_10kt"] or 0),
        reverse=True,
    )

    top_supplier = latest_imports[0]["importer"] if latest_imports else ""
    latest_concentration = float(latest_imports[0]["import_concentration_pct"] or 0) / 100 if latest_imports else 0

    risk = {}
    for row in risk_rows:
        country = row["country"]
        level = row["risk_level"]
        risk[country] = {
            "ci2022": round2(row["ci_2022"]),
            "ci2023": round2(row["ci_2023"]),
            "ci2024": round2(row["ci_2024"]),
            "score": round2(row["ci_mean_3yr"]),
            "level": level,
            "levelCn": row["risk_level_cn"],
            "nYears": int(row["n_years"] or 0),
        }

    supplier_risk = risk.get(top_supplier, {}).get("level", "medium")

    short_term = []
    for row in short_rows:
        country = row["reporter_name"]
        raw_level = row["sub_level"]
        if float(row["sub_potential"] or 0) <= 0:
            continue
        short_term.append(
            {
                "country": country,
                "potential": round2(row["sub_potential"]),
                "exportAvg": round2(float(row["export_avg_kg"] or 0) / 1_000_000_000),
                "importAvg": round2(float(row["import_avg_kg"] or 0) / 1_000_000_000),
                "tier": SHORT_TERM_TIER.get(raw_level, 3),
                "tierLabel": raw_level,
            }
        )
    short_term.sort(key=lambda item: item["potential"], reverse=True)

    long_term = []
    for row in long_rows:
        country = row["country"]
        raw_level = row["sub_tier"]
        ratio = float(row["substitution_ratio"] or 0)
        if ratio <= 0:
            continue
        long_term.append(
            {
                "country": country,
                "potential": round2(ratio),
                "potentialProduction": round2(float(row["potential_prod_1000t"] or 0) / 1000),
                "actualProduction": round2(float(row["actual_prod_1000t"] or 0) / 1000),
                "gap": round2(float(row["gap_1000t"] or 0) / 1000),
                "importAvg": round2(float(row["import_avg_1000t"] or 0) / 1000),
                "tier": LONG_TERM_TIER.get(raw_level, 3),
                "tierLabel": raw_level,
            }
        )
    long_term.sort(key=lambda item: item["potential"], reverse=True)

    return {
        "id": crop_id,
        "name": config["name"],
        "subtitle": config["subtitle"],
        "overview": {
            "latestProduction": production_trend[-1],
            "importConcentration": round2(latest_concentration),
            "topSupplier": top_supplier,
            "riskHeadline": RISK_LABELS.get(supplier_risk, "中风险"),
            "chinaImportTotal": import_trend[-1],
        },
        "years": [str(year) for year in years],
        "productionTrend": production_trend,
        "importTrend": import_trend,
        "producers": producers,
        "imports": imports,
        "risk": risk,
        "alternatives": {
            "shortTerm": short_term[:20],
            "longTerm": long_term[:20],
        },
    }


def build_dataset():
    products = [build_product(crop_id, config) for crop_id, config in CROPS.items()]
    all_years = products[0]["years"]
    return {
        "years": all_years,
        "countryLabels": COUNTRY_LABELS,
        "countryCoordinates": COUNTRY_COORDINATES,
        "products": products,
    }


def validate_data_root(data_root=None):
    data_root = Path(data_root or ROOT)
    missing_dirs = [name for name in REQUIRED_DIRS if not (data_root / name).is_dir()]
    if missing_dirs:
        raise FileNotFoundError(
            "数据目录缺少以下文件夹："
            + "、".join(missing_dirs)
            + f"\n当前 ROOT={data_root}"
        )

    sample_file = data_root / "前10大主产国汇总" / "大豆前10主产国汇总(soybean_top10_producers).xlsx"
    if not sample_file.is_file():
        raise FileNotFoundError(
            "未找到示例 Excel 文件："
            f"\n{sample_file}"
            + "\n请确认数据目录中包含实验室导出的 Excel 结果。"
        )


def main():
    print(f"ROOT={ROOT}")
    print(f"WORKSPACE={WORKSPACE}")
    validate_data_root()
    dataset = build_dataset()
    payload = "window.GRAIN_APP_DATA = " + json.dumps(dataset, ensure_ascii=False, indent=2) + ";\n"
    OUTPUT.write_text(payload, encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
